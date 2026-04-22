#!/usr/bin/env python3
"""
Run share-of-answers baseline across ChatGPT (with web search) and Perplexity APIs.

For each prompt x each engine x N runs:
  * Send prompt to engine
  * Parse cited URLs from response
  * Check whether target domain appears (cite = 1, not cited = 0)
  * Average across runs → per-prompt cite rate per engine

Writes results back to the opportunity-sizing xlsx (02_Prompts_Raw columns J + K)
and saves raw responses to JSONL for audit.

Usage:
    # Dry run to confirm cost estimate
    python3 tools/geo_baseline_runner.py \
        --prompts "clients/vumedi/projects/AEO Audit/opportunity-sizing/oncology-prompts-v1.csv" \
        --xlsx    "clients/vumedi/projects/AEO Audit/opportunity-sizing/geo-opportunity-model-shell.xlsx" \
        --target-domain vumedi.com \
        --dry-run

    # Test on 3 prompts
    python3 tools/geo_baseline_runner.py ... --limit 3

    # Full run
    python3 tools/geo_baseline_runner.py ...

Requires OPENAI_API_KEY and PERPLEXITY_API_KEY in environment or .env.
"""
import argparse
import csv
import json
import os
import sys
import time
from pathlib import Path
from statistics import mean

import requests
from openai import OpenAI
from openpyxl import load_workbook

# Rough cost per query (USD)
COST_PER_CHATGPT = 0.015  # gpt-4o with web_search_preview (low context)
COST_PER_PERPLEXITY = 0.013  # sonar-pro w/ search

OPENAI_MODEL = "gpt-4o"
PERPLEXITY_MODEL = "sonar-pro"
PERPLEXITY_URL = "https://api.perplexity.ai/chat/completions"


def load_env(env_path: Path):
    """Minimal .env loader — avoids adding python-dotenv dependency."""
    if not env_path.exists():
        return
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        v = v.strip().strip('"').strip("'")
        os.environ.setdefault(k.strip(), v)


def load_prompts(csv_path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def has_target_domain(urls, domain):
    domain = domain.lower()
    for u in urls:
        if u and domain in u.lower():
            return True
    return False


def query_chatgpt(client: OpenAI, prompt_text: str) -> dict:
    """Query GPT-4o with web_search_preview. Returns dict with cited_urls + raw."""
    try:
        resp = client.responses.create(
            model=OPENAI_MODEL,
            tools=[{"type": "web_search_preview"}],
            tool_choice={"type": "web_search_preview"},  # force web search so every query gets citations
            input=prompt_text,
        )
    except Exception as e:
        return {"error": str(e), "cited_urls": [], "raw": None}

    cited_urls = []
    retrieved_urls = []
    try:
        for item in resp.output:
            item_type = getattr(item, "type", None)
            if item_type == "message":
                content = getattr(item, "content", []) or []
                for block in content:
                    annotations = getattr(block, "annotations", []) or []
                    for ann in annotations:
                        ann_type = getattr(ann, "type", None)
                        if ann_type == "url_citation":
                            url = getattr(ann, "url", None)
                            if url:
                                cited_urls.append(url)
            elif item_type == "web_search_call":
                # Some versions expose results here; best-effort
                results = getattr(item, "results", None) or []
                for r in results:
                    url = getattr(r, "url", None) if hasattr(r, "url") else r.get("url") if isinstance(r, dict) else None
                    if url:
                        retrieved_urls.append(url)
    except Exception as e:
        return {"error": f"parse_error: {e}", "cited_urls": [], "retrieved_urls": [], "raw": str(resp)[:500]}

    return {
        "cited_urls": cited_urls,
        "retrieved_urls": retrieved_urls,
        "error": None,
    }


def query_perplexity(api_key: str, prompt_text: str) -> dict:
    """Query Perplexity sonar-pro. Returns dict with cited_urls + raw."""
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": PERPLEXITY_MODEL,
        "messages": [{"role": "user", "content": prompt_text}],
    }
    try:
        r = requests.post(PERPLEXITY_URL, headers=headers, json=payload, timeout=60)
        if r.status_code != 200:
            return {"error": f"HTTP {r.status_code}: {r.text[:200]}", "cited_urls": []}
        data = r.json()
    except Exception as e:
        return {"error": str(e), "cited_urls": []}

    # Perplexity returns citations at top level or inside choices
    cited_urls = data.get("citations") or []
    # Newer API puts them in search_results
    if not cited_urls:
        sr = data.get("search_results") or []
        cited_urls = [s.get("url") for s in sr if s.get("url")]

    return {
        "cited_urls": cited_urls,
        "error": None,
    }


def run_single(prompt_row, engines, runs, openai_client, perplexity_key, target_domain,
               raw_writer, progress_prefix):
    """Run one prompt across selected engines × N runs. Returns dict with per-engine cite rates."""
    pid = prompt_row["prompt_id"]
    text = prompt_row["prompt_text"]
    result = {"prompt_id": pid, "prompt_text": text, "engines": {}}

    for eng in engines:
        hits = []
        for run_idx in range(runs):
            if eng == "chatgpt":
                out = query_chatgpt(openai_client, text)
            elif eng == "perplexity":
                out = query_perplexity(perplexity_key, text)
            else:
                continue

            err = out.get("error")
            if err:
                print(f"    [{progress_prefix}] {pid} {eng} run{run_idx+1}: ERROR {err}")
                hits.append(None)  # exclude from average
            else:
                hit = has_target_domain(out.get("cited_urls", []), target_domain)
                hits.append(1 if hit else 0)

            # Write raw to JSONL
            raw_writer({
                "prompt_id": pid,
                "engine": eng,
                "run": run_idx + 1,
                "cited_urls": out.get("cited_urls", []),
                "retrieved_urls": out.get("retrieved_urls", []),
                "error": err,
                "has_target": 1 if (err is None and has_target_domain(out.get("cited_urls", []), target_domain)) else 0,
                "target_domain": target_domain,
                "timestamp": time.time(),
            })

            # Light rate limit safety
            time.sleep(0.3)

        clean = [h for h in hits if h is not None]
        rate = mean(clean) if clean else None
        result["engines"][eng] = {
            "runs": hits,
            "cite_rate": rate,
            "n_clean": len(clean),
        }
        print(f"    [{progress_prefix}] {pid} {eng}: rate={rate} (hits={hits})")

    return result


def write_to_xlsx(xlsx_path, results_by_id, engines):
    """Write cite rates into 02_Prompts_Raw columns J (chatgpt) and K (perplexity)."""
    wb = load_workbook(xlsx_path)
    ws = wb["02_Prompts_Raw"]

    # Build header → column map from row 1
    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            header_to_col[h] = col

    col_chatgpt = header_to_col.get("baseline_cite_rate_chatgpt")
    col_perp = header_to_col.get("baseline_cite_rate_perplexity")

    updated = 0
    for row in range(2, ws.max_row + 1):
        pid = ws.cell(row=row, column=1).value
        if not pid or pid not in results_by_id:
            continue
        res = results_by_id[pid]
        if "chatgpt" in engines and col_chatgpt:
            rate = res["engines"].get("chatgpt", {}).get("cite_rate")
            if rate is not None:
                ws.cell(row=row, column=col_chatgpt, value=round(rate, 3)).number_format = "0.0%"
        if "perplexity" in engines and col_perp:
            rate = res["engines"].get("perplexity", {}).get("cite_rate")
            if rate is not None:
                ws.cell(row=row, column=col_perp, value=round(rate, 3)).number_format = "0.0%"
        updated += 1

    wb.save(xlsx_path)
    return updated


def load_existing_from_xlsx(xlsx_path, engines):
    """Return dict prompt_id → set of engines already filled (for --resume)."""
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb["02_Prompts_Raw"]
    header_to_col = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            header_to_col[h] = col
    col_chatgpt = header_to_col.get("baseline_cite_rate_chatgpt")
    col_perp = header_to_col.get("baseline_cite_rate_perplexity")

    filled = {}
    for row in range(2, ws.max_row + 1):
        pid = ws.cell(row=row, column=1).value
        if not pid:
            continue
        done = set()
        if col_chatgpt and ws.cell(row=row, column=col_chatgpt).value is not None:
            done.add("chatgpt")
        if col_perp and ws.cell(row=row, column=col_perp).value is not None:
            done.add("perplexity")
        filled[pid] = done
    return filled


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--prompts", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--target-domain", required=True, help="e.g. vumedi.com")
    ap.add_argument("--engines", default="chatgpt,perplexity")
    ap.add_argument("--runs", type=int, default=3)
    ap.add_argument("--limit", type=int, default=0, help="0 = no limit")
    ap.add_argument("--raw-out", default=".tmp/geo-baseline-raw.jsonl")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--resume", action="store_true",
                    help="Skip prompts already filled in the xlsx")
    args = ap.parse_args()

    load_env(Path(".env"))

    engines = [e.strip() for e in args.engines.split(",") if e.strip()]
    for eng in engines:
        if eng not in ("chatgpt", "perplexity"):
            print(f"Unknown engine: {eng}", file=sys.stderr)
            sys.exit(2)

    prompts = load_prompts(args.prompts)
    if args.limit > 0:
        prompts = prompts[:args.limit]

    # Resume: skip prompts already filled
    if args.resume:
        existing = load_existing_from_xlsx(args.xlsx, engines)
        before = len(prompts)
        prompts = [p for p in prompts if not set(engines).issubset(existing.get(p["prompt_id"], set()))]
        print(f"--resume: skipped {before - len(prompts)} already-filled prompts")

    total_calls = len(prompts) * len(engines) * args.runs
    est_cost = 0
    if "chatgpt" in engines:
        est_cost += len(prompts) * args.runs * COST_PER_CHATGPT
    if "perplexity" in engines:
        est_cost += len(prompts) * args.runs * COST_PER_PERPLEXITY

    print("=" * 70)
    print(f"GEO Baseline Runner")
    print(f"  Prompts:        {len(prompts)}")
    print(f"  Engines:        {engines}")
    print(f"  Runs/engine:    {args.runs}")
    print(f"  Total calls:    {total_calls}")
    print(f"  Est. cost:      ${est_cost:.2f}")
    print(f"  Target domain:  {args.target_domain}")
    print(f"  Raw log:        {args.raw_out}")
    print("=" * 70)

    if args.dry_run:
        print("DRY RUN — no API calls made.")
        return

    # Init clients
    openai_client = None
    perplexity_key = None
    if "chatgpt" in engines:
        if not os.getenv("OPENAI_API_KEY"):
            print("OPENAI_API_KEY missing in env/.env", file=sys.stderr); sys.exit(2)
        openai_client = OpenAI()
    if "perplexity" in engines:
        perplexity_key = os.getenv("PERPLEXITY_API_KEY")
        if not perplexity_key:
            print("PERPLEXITY_API_KEY missing in env/.env", file=sys.stderr); sys.exit(2)

    # Ensure raw output dir exists
    Path(args.raw_out).parent.mkdir(parents=True, exist_ok=True)
    raw_f = open(args.raw_out, "a", encoding="utf-8")
    def raw_writer(d):
        raw_f.write(json.dumps(d) + "\n"); raw_f.flush()

    results = {}
    start = time.time()
    for i, p in enumerate(prompts, 1):
        prefix = f"{i}/{len(prompts)}"
        print(f"[{prefix}] {p['prompt_id']}  {p['prompt_text'][:60]}")
        try:
            r = run_single(p, engines, args.runs, openai_client, perplexity_key,
                           args.target_domain, raw_writer, prefix)
            results[p["prompt_id"]] = r
        except Exception as e:
            print(f"    FATAL on {p['prompt_id']}: {e}")
            continue

        # Every 20 prompts, incremental xlsx write for safety
        if i % 20 == 0:
            updated = write_to_xlsx(args.xlsx, results, engines)
            elapsed = time.time() - start
            print(f"  [checkpoint] wrote {updated} rows to xlsx, elapsed {elapsed:.0f}s")

    raw_f.close()

    # Final write
    updated = write_to_xlsx(args.xlsx, results, engines)
    total = time.time() - start
    print("=" * 70)
    print(f"Done. Wrote {updated} prompt rows to {args.xlsx}")
    print(f"Raw responses: {args.raw_out}")
    print(f"Elapsed: {total:.0f}s ({total/60:.1f} min)")
    print("=" * 70)

    # Summary stats
    for eng in engines:
        rates = [r["engines"].get(eng, {}).get("cite_rate") for r in results.values()]
        rates = [x for x in rates if x is not None]
        if rates:
            cited = sum(1 for x in rates if x > 0)
            print(f"  {eng}: {cited}/{len(rates)} prompts cited VuMedi (avg cite rate {mean(rates):.1%})")


if __name__ == "__main__":
    main()
