#!/usr/bin/env python3
"""
Build the VuMedi Oncology GEO Opportunity Model spreadsheet shell.

Reads: oncology-prompts-v1.csv (the prompt set)
Writes: geo-opportunity-model-shell.xlsx (5-tab model, formulas wired to Assumptions)

Design principles:
  * Every derived cell is a formula referencing 01_Assumptions — never a hardcoded value.
  * Three scenarios (low/mid/high) — model outputs a RANGE, not a point estimate.
  * Every assumption tagged: measured / benchmarked / hypothesis.
  * Aggressive rounding in Summary — no false precision.
  * Auditable: any output cell → trace back to named input.

Usage:
    python3 tools/geo_opportunity_model.py \
        --prompts "clients/vumedi/projects/AEO Audit/opportunity-sizing/oncology-prompts-v1.csv" \
        --out     "clients/vumedi/projects/AEO Audit/opportunity-sizing/geo-opportunity-model-shell.xlsx"
"""
import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ---------- styling ----------
BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MUTED_FILL = PatternFill("solid", fgColor="F3F4F6")
HYPOTHESIS_FILL = PatternFill("solid", fgColor="FEF3C7")   # amber
MEASURED_FILL = PatternFill("solid", fgColor="D1FAE5")     # green
BENCHMARKED_FILL = PatternFill("solid", fgColor="DBEAFE")  # blue
THIN = Side(style="thin", color="D1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BOX


def conf_fill(conf):
    return {
        "measured": MEASURED_FILL,
        "benchmarked": BENCHMARKED_FILL,
        "hypothesis": HYPOTHESIS_FILL,
    }.get(conf, MUTED_FILL)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- 00_README ----------
def build_readme(wb):
    ws = wb.create_sheet("00_README")
    lines = [
        ("VuMedi Oncology — GEO Opportunity Model (v1 draft)", BOLD),
        ("", None),
        ("Purpose", BOLD),
        ("Give Roman a credible RANGE for how much visibility + traffic VuMedi could capture", None),
        ("by shipping the oncology AEO pilot (Tier 1 + Tier 2 + Tier 3).", None),
        ("", None),
        ("This is a rough first pass, not a forecast.", BOLD),
        ("The model is designed to be wrong precisely — show the shape of the opportunity,", None),
        ("make every assumption inspectable, and let future data tighten the range.", None),
        ("", None),
        ("How to read it", BOLD),
        ("  1. Every output is a low-mid-high band. Anchor on the range, not the midpoint.", None),
        ("  2. Every formula cell references 01_Assumptions. Change an input there; model updates.", None),
        ("  3. Every assumption has a confidence tag:", None),
        ("       GREEN = measured (we have data)", None),
        ("       BLUE  = benchmarked (industry source cited)", None),
        ("       AMBER = hypothesis (our best judgment, validate in pilot)", None),
        ("  4. Summary tab rounds aggressively — any precision below nearest 1K is false.", None),
        ("", None),
        ("Tabs", BOLD),
        ("  00_README        This page", None),
        ("  01_Assumptions   Single source of truth for multipliers + ceilings (edit here)", None),
        ("  02_Prompts_Raw   110 oncology prompts with GSC data where measured", None),
        ("  03_Model         Per-prompt calculations (formulas only, no hardcodes)", None),
        ("  04_Summary       Aggregate by intent bucket, low-mid-high ranges", None),
        ("", None),
        ("Known gaps (Pass 2 will address)", BOLD),
        ("  * baseline_cite_rate columns in 02 are empty until API run (Pass 2)", None),
        ("  * click_to_start, minutes_per_start, ai_referral_lift in 01 are placeholders (need GA4 bridge)", None),
        ("  * 86/110 prompts are hypothesized — for buckets with no GSC-measured data,", None),
        ("    base volume uses an editable hypothesis in 01_Assumptions (amber highlight)", None),
        ("", None),
        ("Method in one equation", BOLD),
        ("  capturable_minutes = prompt_volume × capturable_share × CTR × click_to_start × min/start", None),
        ("  Run for each prompt × each scenario (low/mid/high) × 12 months.", None),
        ("", None),
        ("Fanout note (important)", BOLD),
        ("  v1 measures share-of-answers on SEED prompts only. Real AI retrieval expands", None),
        ("  each prompt into ~3-10 fanouts before citing. v1 includes 4 fanout variants as", None),
        ("  a calibration sample (ONC-002/006/008/010). Pass 2 will tell us whether fanouts", None),
        ("  shift cite rate enough to warrant full fanout expansion in v2.", None),
        ("", None),
        ("Source material", BOLD),
        ("  AEO Audit/stream-1-findings.md (+ addenda)  — GSC data", None),
        ("  AEO Audit/stream-3-findings.md              — LLM citation testing", None),
        ("  AEO Audit/oncology-pilot-recommendation.md  — pilot scope + KPIs", None),
        ("  geo-research/geo-optimization-plan.md       — overall GEO strategy", None),
    ]
    for r, (text, font) in enumerate(lines, 1):
        c = ws.cell(row=r, column=1, value=text)
        if font:
            c.font = font
    set_widths(ws, [95])


# ---------- 01_Assumptions ----------
# Scalar assumptions (single row each) + bucket ceiling table.
# param_id | parameter | low | mid | high | confidence | source | notes
SCALAR_ASSUMPTIONS = [
    ("A01", "llm_volume_multiplier", 0.20, 0.40, 0.80, "benchmarked",
     "Graphite: AI ~34% of US search volume (2026)",
     "Fraction of search volume that also occurs on LLM surfaces. Applied as (1 + mult) uplift to GSC-measured volume."),
    ("A02", "hypothesized_volume_mult", 0.50, 1.00, 2.00, "hypothesis",
     "Judgment",
     "Spread around bucket-median volume for prompts with no GSC data. Low = half median, High = 2x median."),
    ("A03", "serp_ctr_target", 0.010, 0.020, 0.030, "measured",
     "VuMedi GSC: 1.33% today, 2.5% Tier-2 target (Stream 1)",
     "Weighted CTR on SERP impressions post-pilot. Conservative to aggressive."),
    ("A04", "llm_citation_ctr", 0.005, 0.015, 0.025, "benchmarked",
     "Industry AI citation click-through ~1-2%",
     "Click-through when VuMedi is cited in an AI answer. Will be measured in pilot; for now a bench."),
    ("A05", "llm_share_of_surface", 0.25, 0.35, 0.50, "benchmarked",
     "Graphite 34% + growth; conservative low",
     "What share of total queries happen on LLMs vs SERPs. Used to weight SERP vs LLM CTR."),
    ("A06", "click_to_start_ratio", 0.30, 0.45, 0.60, "hypothesis",
     "PLACEHOLDER - needs GA4 bridge (Stream 1 addendum flagged)",
     "Organic landing → video start. Using rough benchmark until GA4 view instrumented."),
    ("A07", "minutes_per_start", 4.0, 8.0, 14.0, "hypothesis",
     "PLACEHOLDER - needs GA4 bridge",
     "Minutes watched per video start. Needs real distribution; single value is directional only."),
    ("A08", "ai_referral_lift", 1.0, 2.0, 4.0, "benchmarked",
     "Graphite: 6x claim; conservative for v1",
     "Conversion lift on AI-referred vs search-referred traffic. Applied to click→start→minutes."),
]

# Bucket capturable-share ceilings + hypothesis volume fallback.
# hyp_vol_monthly is used when the bucket has NO gsc_measured prompts to average from.
# v2: Revised after Pass 2 baseline run (1/110 prompts cited — see pass2-findings.md).
BUCKET_CEILINGS = [
    # (bucket_key, label, low, mid, high, confidence, rationale, hyp_vol_monthly)
    ("trial", "Trial reference (KEYNOTE, NATALEE, etc.)",
     0.05, 0.15, 0.30, "hypothesis",
     "Pass 2: 0/60 cite rate. NEJM/OncLive/ASCO/pharma-HCP deeply entrenched. Tier-2 gets us into top 10, not dominant.",
     None),
    ("guideline_classification", "Guidelines + classifications (NCCN, WHO, AJCC)",
     0.03, 0.07, 0.15, "hypothesis",
     "Pass 2: 0/39 cite rate. NCCN/WHO/AJCC are primary sources by definition. Almost no path here.",
     400),
    ("drug_mechanism", "Drug + mechanism + comparison",
     0.10, 0.20, 0.35, "hypothesis",
     "Pass 2: 0/60 cite rate. Pharma HCP sites (keytrudahcp, enhertuhcp) well established — not open. Physician-perspective on comparisons still open.",
     300),
    ("technique", "Surgical + procedural technique",
     0.15, 0.30, 0.50, "hypothesis",
     "Pass 2: 0/45 cite rate. Oncology techniques != orthopedic techniques competitively. Downward revision from v1.",
     200),
    ("category_survey", "Category survey / platform-preference",
     0.15, 0.30, 0.50, "hypothesis",
     "Pass 2: 5.6% cite rate on ChatGPT (the ONE bright spot). Entity recognition exists; authority work compounds.",
     50),
    ("conference_society", "Conference + society",
     0.10, 0.20, 0.40, "hypothesis",
     "Pass 2: 0/44 cite rate. Societies own their own conference queries. Our angle is content coverage, not brand.",
     None),
    ("institution_kol", "Institution + KOL",
     0.05, 0.15, 0.30, "hypothesis",
     "Pass 2: 0/38 cite rate. MSK/Dana-Farber/MDACC dominate their own brand queries. Need hub pages as 'watch [institution] content' surface.",
     None),
]


def build_assumptions(wb):
    ws = wb.create_sheet("01_Assumptions")

    # Top-of-sheet instruction
    ws["A1"] = "ASSUMPTIONS — edit any low/mid/high value and the entire model updates."
    ws["A1"].font = BOLD
    ws.merge_cells("A1:H1")
    ws["A2"] = "Confidence color code:  GREEN measured  |  BLUE benchmarked  |  AMBER hypothesis"
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:H2")

    # Scalar assumptions block
    ws["A4"] = "SCALAR ASSUMPTIONS"
    ws["A4"].font = BOLD
    header = ["param_id", "parameter", "low", "mid", "high", "confidence", "source", "notes"]
    for col, h in enumerate(header, 1):
        ws.cell(row=5, column=col, value=h)
    style_header([ws.cell(row=5, column=c) for c in range(1, len(header) + 1)])

    row = 6
    scalar_row_map = {}  # param_id -> row
    for pid, name, lo, mid, hi, conf, src, notes in SCALAR_ASSUMPTIONS:
        ws.cell(row=row, column=1, value=pid).font = BOLD
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=lo)
        ws.cell(row=row, column=4, value=mid)
        ws.cell(row=row, column=5, value=hi)
        c = ws.cell(row=row, column=6, value=conf)
        c.fill = conf_fill(conf)
        ws.cell(row=row, column=7, value=src)
        ws.cell(row=row, column=8, value=notes)
        scalar_row_map[pid] = row
        row += 1

    # Bucket capturable-share ceilings
    row += 2
    ws.cell(row=row, column=1, value="BUCKET CAPTURABLE-SHARE CEILINGS").font = BOLD
    row += 1
    bucket_header = ["bucket_key", "label", "low", "mid", "high", "confidence", "rationale"]
    for col, h in enumerate(bucket_header, 1):
        ws.cell(row=row, column=col, value=h)
    style_header([ws.cell(row=row, column=c) for c in range(1, len(bucket_header) + 1)])
    bucket_header_row = row

    row += 1
    bucket_table_start = row
    for key, label, lo, mid, hi, conf, rationale, _hyp_vol in BUCKET_CEILINGS:
        ws.cell(row=row, column=1, value=key).font = BOLD
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=3, value=lo)
        ws.cell(row=row, column=4, value=mid)
        ws.cell(row=row, column=5, value=hi)
        c = ws.cell(row=row, column=6, value=conf)
        c.fill = conf_fill(conf)
        ws.cell(row=row, column=7, value=rationale)
        row += 1
    bucket_table_end = row - 1

    # Bucket base monthly volume — GSC avg if measured data exists, else editable hypothesis fallback.
    # Placed below bucket ceilings so formulas can reference it.
    row += 2
    ws.cell(row=row, column=1, value="BUCKET BASE MONTHLY VOLUME (used for prompts without GSC-measured data)").font = BOLD
    row += 1
    ws.cell(row=row, column=1,
            value="If a bucket has GSC-measured prompts, we use their avg. If not, we use the hypothesis (EDITABLE)."
            ).font = Font(italic=True, color="6B7280")
    row += 1
    med_header = ["bucket_key", "avg_monthly_imps_gsc", "count_gsc_measured",
                  "hyp_monthly_imps", "base_volume_used", "notes"]
    for col, h in enumerate(med_header, 1):
        ws.cell(row=row, column=col, value=h)
    style_header([ws.cell(row=row, column=c) for c in range(1, len(med_header) + 1)])
    median_header_row = row

    row += 1
    median_table_start = row
    bucket_median_row_map = {}
    for key, _label, _lo, _mid, _hi, _conf, _rat, hyp_vol in BUCKET_CEILINGS:
        ws.cell(row=row, column=1, value=key).font = BOLD
        # Column B: avg monthly imps across GSC-measured rows for this bucket
        avg_formula = (
            f'=IFERROR(AVERAGEIFS(\'02_Prompts_Raw\'!E:E,'
            f'\'02_Prompts_Raw\'!B:B,"{key}",'
            f'\'02_Prompts_Raw\'!D:D,"gsc_measured")/16,0)'
        )
        cell = ws.cell(row=row, column=2, value=avg_formula)
        cell.number_format = "#,##0"
        # Column C: count of GSC-measured prompts in this bucket
        count_formula = (
            f'=COUNTIFS(\'02_Prompts_Raw\'!B:B,"{key}",'
            f'\'02_Prompts_Raw\'!D:D,"gsc_measured")'
        )
        ws.cell(row=row, column=3, value=count_formula)
        # Column D: hypothesis fallback volume (editable, only populated for buckets without GSC data)
        hyp_cell = ws.cell(row=row, column=4, value=hyp_vol if hyp_vol else "")
        hyp_cell.number_format = "#,##0"
        if hyp_vol:
            hyp_cell.fill = HYPOTHESIS_FILL
        # Column E: base_volume_used — picks GSC avg if count>0, else hypothesis
        base_formula = f"=IF(C{row}>0,B{row},D{row})"
        cell_e = ws.cell(row=row, column=5, value=base_formula)
        cell_e.number_format = "#,##0"
        cell_e.font = BOLD
        # Column F: notes
        note = ("GSC-derived (has measured data)"
                if hyp_vol is None
                else "HYPOTHESIS — no GSC data for bucket; estimate based on query-intent pattern")
        ws.cell(row=row, column=6, value=note)
        bucket_median_row_map[key] = row
        row += 1
    median_table_end = row - 1

    set_widths(ws, [14, 40, 10, 10, 10, 14, 50, 70])

    # Return references so other sheets can build formulas
    return {
        "scalar_row_map": scalar_row_map,  # "A01" -> row number
        "bucket_ceiling_header_row": bucket_header_row,
        "bucket_ceiling_start": bucket_table_start,
        "bucket_ceiling_end": bucket_table_end,
        "bucket_median_header_row": median_header_row,
        "bucket_median_start": median_table_start,
        "bucket_median_end": median_table_end,
        "bucket_median_row_map": bucket_median_row_map,
    }


# ---------- 02_Prompts_Raw ----------
def read_existing_baselines(xlsx_path):
    """If the xlsx already exists, harvest baseline_cite_rate values so we don't
    overwrite them when regenerating after an assumption tweak."""
    if not xlsx_path or not Path(xlsx_path).exists():
        return {}
    try:
        from openpyxl import load_workbook as _lwb
        wb = _lwb(xlsx_path, data_only=False)
        if "02_Prompts_Raw" not in wb.sheetnames:
            return {}
        ws = wb["02_Prompts_Raw"]
        header_to_col = {ws.cell(row=1, column=c).value: c
                         for c in range(1, ws.max_column + 1)
                         if ws.cell(row=1, column=c).value}
        col_cgpt = header_to_col.get("baseline_cite_rate_chatgpt")
        col_perp = header_to_col.get("baseline_cite_rate_perplexity")
        out = {}
        for r in range(2, ws.max_row + 1):
            pid = ws.cell(row=r, column=1).value
            if not pid:
                continue
            out[pid] = (
                ws.cell(row=r, column=col_cgpt).value if col_cgpt else None,
                ws.cell(row=r, column=col_perp).value if col_perp else None,
            )
        return out
    except Exception:
        return {}


def build_prompts_raw(wb, prompts_csv, existing_baselines=None):
    existing_baselines = existing_baselines or {}
    ws = wb.create_sheet("02_Prompts_Raw")
    header = [
        "prompt_id", "bucket", "prompt_text", "source",
        "gsc_16mo_impressions", "gsc_avg_position",
        "has_vumedi_content", "tier2_relevant", "notes",
        "baseline_cite_rate_chatgpt", "baseline_cite_rate_perplexity", "baseline_cite_rate_avg",
    ]
    for col, h in enumerate(header, 1):
        ws.cell(row=1, column=col, value=h)
    style_header([ws.cell(row=1, column=c) for c in range(1, len(header) + 1)])

    with open(prompts_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    for r, row in enumerate(rows, 2):
        pid = row["prompt_id"]
        ws.cell(row=r, column=1, value=pid)
        ws.cell(row=r, column=2, value=row["bucket"])
        ws.cell(row=r, column=3, value=row["prompt_text"])
        ws.cell(row=r, column=4, value=row["source"])
        gsc_imps = row.get("gsc_16mo_impressions", "").strip()
        ws.cell(row=r, column=5, value=int(gsc_imps) if gsc_imps else None)
        pos = row.get("gsc_avg_position", "").strip()
        ws.cell(row=r, column=6, value=float(pos) if pos else None)
        ws.cell(row=r, column=7, value=row.get("has_vumedi_content", ""))
        ws.cell(row=r, column=8, value=row.get("tier2_relevant", ""))
        ws.cell(row=r, column=9, value=row.get("notes", ""))
        # Baseline cite rates — preserve existing if we have them
        cgpt, perp = existing_baselines.get(pid, (None, None))
        cell_j = ws.cell(row=r, column=10, value=cgpt)
        cell_k = ws.cell(row=r, column=11, value=perp)
        if cgpt is not None:
            cell_j.number_format = "0.0%"
        if perp is not None:
            cell_k.number_format = "0.0%"
        avg_formula = f"=IFERROR(AVERAGE(J{r}:K{r}),\"\")"
        cell_l = ws.cell(row=r, column=12, value=avg_formula)
        cell_l.number_format = "0.0%"

    set_widths(ws, [10, 22, 45, 16, 12, 12, 12, 12, 40, 14, 14, 14])
    ws.freeze_panes = "A2"

    return {"data_rows": len(rows), "last_row": len(rows) + 1}


# ---------- 03_Model ----------
def build_model(wb, prompt_count, last_prompt_row, assumptions_refs):
    ws = wb.create_sheet("03_Model")

    # Column layout
    cols = [
        ("prompt_id", 10),
        ("bucket", 22),
        ("source", 14),
        ("gsc_16mo_imps", 12),
        ("monthly_search_imps_LOW", 14),
        ("monthly_search_imps_MID", 14),
        ("monthly_search_imps_HIGH", 14),
        ("total_monthly_queries_LOW", 16),
        ("total_monthly_queries_MID", 16),
        ("total_monthly_queries_HIGH", 16),
        ("capturable_share_LOW", 14),
        ("capturable_share_MID", 14),
        ("capturable_share_HIGH", 14),
        ("annual_capturable_imps_LOW", 18),
        ("annual_capturable_imps_MID", 18),
        ("annual_capturable_imps_HIGH", 18),
        ("annual_capturable_clicks_LOW", 18),
        ("annual_capturable_clicks_MID", 18),
        ("annual_capturable_clicks_HIGH", 18),
        ("annual_capturable_minutes_LOW", 18),
        ("annual_capturable_minutes_MID", 18),
        ("annual_capturable_minutes_HIGH", 18),
    ]
    for col, (h, _w) in enumerate(cols, 1):
        ws.cell(row=1, column=col, value=h)
    style_header([ws.cell(row=1, column=c) for c in range(1, len(cols) + 1)])
    set_widths(ws, [w for _h, w in cols])
    ws.freeze_panes = "D2"

    # Shortcut refs to 01_Assumptions rows
    sr = assumptions_refs["scalar_row_map"]
    a_llm_lo = f"'01_Assumptions'!C{sr['A01']}"
    a_llm_mid = f"'01_Assumptions'!D{sr['A01']}"
    a_llm_hi = f"'01_Assumptions'!E{sr['A01']}"
    a_hyp_lo = f"'01_Assumptions'!C{sr['A02']}"
    a_hyp_mid = f"'01_Assumptions'!D{sr['A02']}"
    a_hyp_hi = f"'01_Assumptions'!E{sr['A02']}"
    a_serp_lo = f"'01_Assumptions'!C{sr['A03']}"
    a_serp_mid = f"'01_Assumptions'!D{sr['A03']}"
    a_serp_hi = f"'01_Assumptions'!E{sr['A03']}"
    a_llmctr_lo = f"'01_Assumptions'!C{sr['A04']}"
    a_llmctr_mid = f"'01_Assumptions'!D{sr['A04']}"
    a_llmctr_hi = f"'01_Assumptions'!E{sr['A04']}"
    a_llmsurface_lo = f"'01_Assumptions'!C{sr['A05']}"
    a_llmsurface_mid = f"'01_Assumptions'!D{sr['A05']}"
    a_llmsurface_hi = f"'01_Assumptions'!E{sr['A05']}"
    a_c2s_lo = f"'01_Assumptions'!C{sr['A06']}"
    a_c2s_mid = f"'01_Assumptions'!D{sr['A06']}"
    a_c2s_hi = f"'01_Assumptions'!E{sr['A06']}"
    a_mps_lo = f"'01_Assumptions'!C{sr['A07']}"
    a_mps_mid = f"'01_Assumptions'!D{sr['A07']}"
    a_mps_hi = f"'01_Assumptions'!E{sr['A07']}"
    a_lift_lo = f"'01_Assumptions'!C{sr['A08']}"
    a_lift_mid = f"'01_Assumptions'!D{sr['A08']}"
    a_lift_hi = f"'01_Assumptions'!E{sr['A08']}"

    # Bucket tables as ranges for VLOOKUP
    bc_start = assumptions_refs["bucket_ceiling_start"]
    bc_end = assumptions_refs["bucket_ceiling_end"]
    bc_range = f"'01_Assumptions'!$A${bc_start}:$E${bc_end}"
    # Columns in bucket ceiling table: A bucket_key, B label, C low, D mid, E high

    bm_start = assumptions_refs["bucket_median_start"]
    bm_end = assumptions_refs["bucket_median_end"]
    bm_range = f"'01_Assumptions'!$A${bm_start}:$E${bm_end}"
    # Columns: A bucket_key, B avg_monthly_imps_gsc, C count, D hyp, E base_volume_used
    # We VLOOKUP column 5 (base_volume_used) so the fallback is automatic.

    for pr in range(2, last_prompt_row + 1):
        src_col = f"'02_Prompts_Raw'!D{pr}"
        bucket_col = f"'02_Prompts_Raw'!B{pr}"
        gsc_col = f"'02_Prompts_Raw'!E{pr}"

        # Passthrough ID + metadata
        ws.cell(row=pr, column=1, value=f"='02_Prompts_Raw'!A{pr}")
        ws.cell(row=pr, column=2, value=f"='02_Prompts_Raw'!B{pr}")
        ws.cell(row=pr, column=3, value=f"='02_Prompts_Raw'!D{pr}")
        ws.cell(row=pr, column=4, value=f"='02_Prompts_Raw'!E{pr}")

        # monthly_search_imps for each scenario:
        # If gsc_measured: gsc_16mo / 16
        # If hypothesized: bucket_median × hypothesized_volume_mult[scenario]
        # If gsc_fanout: bucket_median × 0.5 × hypothesized_volume_mult[scenario] (fanouts get half)
        # If stream3_tested: bucket_median × hypothesized_volume_mult[scenario]
        def monthly_search(scenario_hyp_mult):
            # Column 5 of bm_range = base_volume_used (GSC avg or hypothesis fallback)
            return (
                f'=IF({src_col}="gsc_measured",{gsc_col}/16,'
                f'IF({src_col}="gsc_fanout",'
                f'VLOOKUP({bucket_col},{bm_range},5,FALSE)*0.5*{scenario_hyp_mult},'
                f'VLOOKUP({bucket_col},{bm_range},5,FALSE)*{scenario_hyp_mult}))'
            )
        ws.cell(row=pr, column=5, value=monthly_search(a_hyp_lo))
        ws.cell(row=pr, column=6, value=monthly_search(a_hyp_mid))
        ws.cell(row=pr, column=7, value=monthly_search(a_hyp_hi))

        # total_monthly_queries = monthly_search × (1 + llm_multiplier[scenario])
        ws.cell(row=pr, column=8, value=f"=E{pr}*(1+{a_llm_lo})")
        ws.cell(row=pr, column=9, value=f"=F{pr}*(1+{a_llm_mid})")
        ws.cell(row=pr, column=10, value=f"=G{pr}*(1+{a_llm_hi})")

        # capturable_share lookup from bucket ceiling table
        # Column 3 = low, 4 = mid, 5 = high
        ws.cell(row=pr, column=11, value=f"=VLOOKUP({bucket_col},{bc_range},3,FALSE)")
        ws.cell(row=pr, column=12, value=f"=VLOOKUP({bucket_col},{bc_range},4,FALSE)")
        ws.cell(row=pr, column=13, value=f"=VLOOKUP({bucket_col},{bc_range},5,FALSE)")

        # annual_capturable_imps = total_monthly × share × 12
        ws.cell(row=pr, column=14, value=f"=H{pr}*K{pr}*12")
        ws.cell(row=pr, column=15, value=f"=I{pr}*L{pr}*12")
        ws.cell(row=pr, column=16, value=f"=J{pr}*M{pr}*12")

        # Weighted CTR:
        #   weighted_ctr = (1 - llm_share) * serp_ctr + llm_share * llm_ctr
        def weighted_ctr(llm_share, serp_ctr, llm_ctr):
            return f"((1-{llm_share})*{serp_ctr}+{llm_share}*{llm_ctr})"
        ctr_lo = weighted_ctr(a_llmsurface_lo, a_serp_lo, a_llmctr_lo)
        ctr_mid = weighted_ctr(a_llmsurface_mid, a_serp_mid, a_llmctr_mid)
        ctr_hi = weighted_ctr(a_llmsurface_hi, a_serp_hi, a_llmctr_hi)

        # annual_capturable_clicks = annual_capturable_imps × weighted_ctr
        ws.cell(row=pr, column=17, value=f"=N{pr}*{ctr_lo}")
        ws.cell(row=pr, column=18, value=f"=O{pr}*{ctr_mid}")
        ws.cell(row=pr, column=19, value=f"=P{pr}*{ctr_hi}")

        # annual_capturable_minutes = clicks × click_to_start × minutes_per_start × ai_referral_lift
        ws.cell(row=pr, column=20, value=f"=Q{pr}*{a_c2s_lo}*{a_mps_lo}*{a_lift_lo}")
        ws.cell(row=pr, column=21, value=f"=R{pr}*{a_c2s_mid}*{a_mps_mid}*{a_lift_mid}")
        ws.cell(row=pr, column=22, value=f"=S{pr}*{a_c2s_hi}*{a_mps_hi}*{a_lift_hi}")

        # Number formatting
        for col in (5, 6, 7, 8, 9, 10):
            ws.cell(row=pr, column=col).number_format = "#,##0"
        for col in (11, 12, 13):
            ws.cell(row=pr, column=col).number_format = "0.0%"
        for col in (14, 15, 16, 17, 18, 19, 20, 21, 22):
            ws.cell(row=pr, column=col).number_format = "#,##0"

    return {"last_row": last_prompt_row}


# ---------- 04_Summary ----------
def build_summary(wb, last_prompt_row, buckets):
    ws = wb.create_sheet("04_Summary")

    # Round to nearest 1K for impressions, nearest 100 for clicks, nearest 1K for minutes
    ws["A1"] = "SUMMARY — annualized ranges, oncology AEO pilot"
    ws["A1"].font = BOLD
    ws.merge_cells("A1:H1")
    ws["A2"] = "All values rounded to prevent false precision. Ranges are low/mid/high scenarios."
    ws["A2"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A2:H2")

    # By-bucket table
    ws["A4"] = "BY BUCKET — annual capturable minutes (rounded to nearest 1,000)"
    ws["A4"].font = BOLD
    header = ["bucket", "# prompts", "capturable_imps_LOW", "capturable_imps_MID", "capturable_imps_HIGH",
              "capturable_clicks_LOW", "capturable_clicks_MID", "capturable_clicks_HIGH",
              "capturable_minutes_LOW", "capturable_minutes_MID", "capturable_minutes_HIGH"]
    for col, h in enumerate(header, 1):
        ws.cell(row=5, column=col, value=h)
    style_header([ws.cell(row=5, column=c) for c in range(1, len(header) + 1)])

    row = 6
    for bucket_key in buckets:
        ws.cell(row=row, column=1, value=bucket_key).font = BOLD
        ws.cell(row=row, column=2,
                value=f'=COUNTIF(\'02_Prompts_Raw\'!B:B,"{bucket_key}")')
        # Capturable imps — sum from 03_Model (cols N/O/P = 14/15/16), rounded to nearest 1K
        for i, col_letter in enumerate(("N", "O", "P"), start=3):
            f = (
                f'=MROUND(SUMIF(\'03_Model\'!B:B,"{bucket_key}",\'03_Model\'!{col_letter}:{col_letter}),1000)'
            )
            c = ws.cell(row=row, column=i, value=f)
            c.number_format = "#,##0"
        # Clicks — Q/R/S = 17/18/19, rounded to nearest 100
        for i, col_letter in enumerate(("Q", "R", "S"), start=6):
            f = (
                f'=MROUND(SUMIF(\'03_Model\'!B:B,"{bucket_key}",\'03_Model\'!{col_letter}:{col_letter}),100)'
            )
            c = ws.cell(row=row, column=i, value=f)
            c.number_format = "#,##0"
        # Minutes — T/U/V = 20/21/22, rounded to nearest 1K
        for i, col_letter in enumerate(("T", "U", "V"), start=9):
            f = (
                f'=MROUND(SUMIF(\'03_Model\'!B:B,"{bucket_key}",\'03_Model\'!{col_letter}:{col_letter}),1000)'
            )
            c = ws.cell(row=row, column=i, value=f)
            c.number_format = "#,##0"
        row += 1

    # Grand total row
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL (oncology)").font = BOLD
    ws.cell(row=total_row, column=2, value=f"=SUM(B6:B{row - 1})")
    for col in range(3, 12):
        col_letter = get_column_letter(col)
        ws.cell(row=total_row, column=col,
                value=f"=SUM({col_letter}6:{col_letter}{row - 1})").number_format = "#,##0"
    for col in range(1, 12):
        ws.cell(row=total_row, column=col).fill = MUTED_FILL
        ws.cell(row=total_row, column=col).font = BOLD

    # Headline callout
    row += 3
    ws.cell(row=row, column=1, value="HEADLINE RANGES (round for exec memo)").font = BOLD
    row += 2
    ws.cell(row=row, column=1, value="Annual capturable impressions:").font = BOLD
    ws.cell(row=row, column=3, value=f'=ROUND(C{total_row}/1000000,1)&"M – "&ROUND(E{total_row}/1000000,1)&"M"')
    row += 1
    ws.cell(row=row, column=1, value="Annual capturable clicks:").font = BOLD
    ws.cell(row=row, column=3, value=f'=ROUND(F{total_row}/1000,0)&"K – "&ROUND(H{total_row}/1000,0)&"K"')
    row += 1
    ws.cell(row=row, column=1, value="Annual capturable minutes:").font = BOLD
    ws.cell(row=row, column=3, value=f'=ROUND(I{total_row}/1000,0)&"K – "&ROUND(K{total_row}/1000,0)&"K"')

    # Sensitivity guidance
    row += 3
    ws.cell(row=row, column=1, value="WHAT WOULD TIGHTEN THE RANGE").font = BOLD
    row += 1
    tightening = [
        "1. GA4 bridge: instrument organic landing → video start → minutes watched per specialty.",
        "2. Ahrefs/Semrush access: replaces 'hypothesized bucket-median' for 86 of 110 prompts.",
        "3. API-driven baseline cite rates (Pass 2): lets us move from 'capturable share' ceiling to 'headroom from current' — tighter range.",
        "4. Category Lead validation of bucket taxonomy + ceilings — captures domain knowledge we don't have.",
    ]
    for t in tightening:
        ws.cell(row=row, column=1, value=t)
        row += 1

    # Top assumptions driving range
    row += 2
    ws.cell(row=row, column=1, value="TOP 3 ASSUMPTIONS DRIVING THE RANGE (edit in 01_Assumptions)").font = BOLD
    row += 1
    top_assumptions = [
        "A02 hypothesized_volume_mult — 0.5x to 2x spread on 86 of 110 prompts.",
        "A01 llm_volume_multiplier — 0.2 to 0.8 LLM-volume uplift beyond search.",
        "Bucket ceilings (esp. trial, category_survey) — hypothesis until Pass 2 baseline measured.",
    ]
    for t in top_assumptions:
        ws.cell(row=row, column=1, value=t)
        row += 1

    set_widths(ws, [28, 10, 16, 16, 16, 16, 16, 16, 16, 16, 16])


# ---------- main ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--prompts", required=True, help="Path to oncology-prompts-v1.csv")
    ap.add_argument("--out", required=True, help="Path to output xlsx")
    args = ap.parse_args()

    # Preserve existing baseline cite rates if the xlsx already exists
    existing_baselines = read_existing_baselines(args.out)
    if existing_baselines:
        n_with_data = sum(1 for v in existing_baselines.values() if any(x is not None for x in v))
        print(f"Preserving {n_with_data} existing baseline_cite_rate rows from prior build")

    wb = Workbook()
    # Remove the default sheet
    default = wb.active
    wb.remove(default)

    build_readme(wb)
    assumptions_refs = build_assumptions(wb)
    raw_refs = build_prompts_raw(wb, args.prompts, existing_baselines=existing_baselines)
    build_model(wb, raw_refs["data_rows"], raw_refs["last_row"], assumptions_refs)
    build_summary(wb, raw_refs["last_row"], [b[0] for b in BUCKET_CEILINGS])

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")
    print(f"  Prompts: {raw_refs['data_rows']}")
    print(f"  Tabs: 00_README, 01_Assumptions, 02_Prompts_Raw, 03_Model, 04_Summary")


if __name__ == "__main__":
    main()
